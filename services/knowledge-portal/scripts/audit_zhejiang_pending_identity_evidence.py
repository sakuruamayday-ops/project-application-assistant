#!/usr/bin/env python3
"""Reconcile pending Zhejiang identities with archived batch-result evidence.

The script is deliberately read-only for both the production index and the
archived workbooks.  It emits an auditable candidate queue; it never promotes a
commercial-source result to ``knowledge_verified`` by itself.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable


USCC_RE = re.compile(r"^[0-9A-Z]{18}$")
NAME_COLUMNS = (
    "导入名称",
    "输入企业名称",
    "名单企业名称",
    "企业名称",
    "当前名称",
    "企知道当前名称",
    "主体名称",
)
EXEMPT_PROJECTS = {"浙江制造精品", "地方科技小巨人企业"}


def normalize_name(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\s\u200b\ufeff]+", "", text)
    return text


def normalize_uscc(value: object) -> str:
    text = re.sub(r"[^0-9A-Za-z]", "", str(value or "")).upper()
    return text if USCC_RE.fullmatch(text) else ""


def json_list(value: object) -> list[str]:
    try:
        parsed = json.loads(str(value or "[]"))
    except json.JSONDecodeError:
        return []
    return [str(item).strip() for item in parsed if str(item).strip()] if isinstance(parsed, list) else []


def iter_result_tables(inspect_path: Path) -> Iterable[tuple[str, list[list[Any]]]]:
    with inspect_path.open("r", encoding="utf-8") as handle:
        for raw_line in handle:
            try:
                item = json.loads(raw_line)
            except json.JSONDecodeError:
                continue
            if item.get("kind") != "table":
                continue
            values = item.get("values")
            if not isinstance(values, list) or not values:
                continue
            headers = [str(value or "").strip() for value in values[0]]
            if "统一社会信用代码" not in headers:
                continue
            yield str(item.get("sheet") or ""), values


def add_candidate_rows(
    *,
    by_name: dict[str, list[dict[str, str]]],
    seen_records: set[tuple[str, str, str, str]],
    values: list[list[Any]],
    source_file: Path,
    source_sheet: str,
) -> int:
    headers = [str(value or "").strip() for value in values[0]]
    positions = {header: index for index, header in enumerate(headers) if header}
    if "统一社会信用代码" not in positions:
        return 0
    code_index = positions["统一社会信用代码"]
    available_names = [column for column in NAME_COLUMNS if column in positions]
    if not available_names:
        return 0
    valid_rows = 0
    for row in values[1:]:
        if not isinstance(row, (list, tuple)):
            continue
        code = normalize_uscc(row[code_index] if code_index < len(row) else "")
        if not code:
            continue
        valid_rows += 1
        values_by_column: dict[str, str] = {}
        for column in available_names:
            index = positions[column]
            value = str(row[index] or "").strip() if index < len(row) else ""
            if value:
                values_by_column[column] = value
        current_name = values_by_column.get("企业名称") or values_by_column.get("当前名称") or values_by_column.get("企知道当前名称") or ""
        import_name = values_by_column.get("导入名称") or values_by_column.get("输入企业名称") or values_by_column.get("名单企业名称") or ""
        for column, name in values_by_column.items():
            normalized = normalize_name(name)
            if not normalized:
                continue
            dedupe_key = (normalized, code, str(source_file), source_sheet)
            if dedupe_key in seen_records:
                continue
            seen_records.add(dedupe_key)
            by_name[normalized].append(
                {
                    "matched_value": name,
                    "matched_column": column,
                    "unified_social_credit_code": code,
                    "current_name": current_name,
                    "import_name": import_name,
                    "source_file": str(source_file),
                    "source_sheet": source_sheet,
                    "evidence_status": "commercial_batch_candidate_pending_independent_verification",
                }
            )
    return valid_rows


def iter_raw_xlsx_tables(raw_xlsx_dir: Path) -> Iterable[tuple[Path, str, list[list[Any]]]]:
    from openpyxl import load_workbook

    for path in sorted(raw_xlsx_dir.glob("批量查询*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                # The vendor exports an invalid A1-only dimension even though the
                # sheet contains a full table.  Recalculate before iterating.
                worksheet.reset_dimensions()
                worksheet.calculate_dimension(force=True)
                rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
                header_index = next(
                    (
                        index
                        for index, row in enumerate(rows[:10])
                        if "统一社会信用代码" in {str(value or "").strip() for value in row}
                    ),
                    None,
                )
                if header_index is None:
                    continue
                yield path, worksheet.title, rows[header_index:]
        finally:
            workbook.close()


def load_archived_candidates(
    inspect_dir: Path, raw_xlsx_dir: Path | None
) -> tuple[dict[str, list[dict[str, str]]], dict[str, Any]]:
    by_name: dict[str, list[dict[str, str]]] = defaultdict(list)
    seen_records: set[tuple[str, str, str, str]] = set()
    scanned_files = 0
    scanned_rows = 0
    raw_files_scanned = 0

    for inspect_path in sorted(inspect_dir.glob("*.xlsx.inspect.ndjson")):
        if "上传名单" in inspect_path.name:
            continue
        scanned_files += 1
        for sheet, values in iter_result_tables(inspect_path):
            scanned_rows += add_candidate_rows(
                by_name=by_name,
                seen_records=seen_records,
                values=values,
                source_file=inspect_path,
                source_sheet=sheet,
            )

    if raw_xlsx_dir:
        seen_raw_files: set[Path] = set()
        for path, sheet, values in iter_raw_xlsx_tables(raw_xlsx_dir):
            seen_raw_files.add(path)
            scanned_rows += add_candidate_rows(
                by_name=by_name,
                seen_records=seen_records,
                values=values,
                source_file=path,
                source_sheet=sheet,
            )
        raw_files_scanned = len(seen_raw_files)

    return by_name, {
        "inspect_files_scanned": scanned_files,
        "raw_xlsx_files_scanned": raw_files_scanned,
        "valid_result_rows_scanned_before_deduplication": scanned_rows,
        "normalized_names_indexed": len(by_name),
        "name_code_source_records": len(seen_records),
    }


def load_pending_profiles(database: Path) -> list[dict[str, Any]]:
    connection = sqlite3.connect(f"file:{database}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            """
            SELECT identity_key,current_name,recognition_names_json,
                   recognition_projects_json,verification_status
            FROM enterprise_identity_profiles
            WHERE verification_status='pending_business_identity'
            ORDER BY current_name,identity_key
            """
        ).fetchall()
    finally:
        connection.close()
    return [dict(row) for row in rows]


def verified_checklist_codes(path: Path | None) -> set[str]:
    if not path:
        return set()
    codes: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        if not re.match(r"^\|\s*\d+\s*\|", line):
            continue
        cells = [cell.strip().strip("`") for cell in line.strip().strip("|").split("|")]
        if len(cells) != 8:
            continue
        code = normalize_uscc(cells[3])
        conclusion = cells[6]
        if code and ("核验通过" in conclusion or "归并" in conclusion):
            codes.add(code)
    return codes


def write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def reconcile(
    database: Path,
    inspect_dir: Path,
    raw_xlsx_dir: Path | None,
    verified_checklist: Path | None,
    output_dir: Path,
) -> dict[str, Any]:
    archived_by_name, archive_stats = load_archived_candidates(inspect_dir, raw_xlsx_dir)
    pending_profiles = load_pending_profiles(database)
    output_dir.mkdir(parents=True, exist_ok=True)

    unique_candidates: list[dict[str, object]] = []
    conflicts: list[dict[str, object]] = []
    missing: list[dict[str, object]] = []
    exempt_only: list[dict[str, object]] = []
    coverage_by_project: Counter[str] = Counter()
    missing_by_project: Counter[str] = Counter()
    conflict_by_project: Counter[str] = Counter()
    verified_codes = verified_checklist_codes(verified_checklist)

    for profile in pending_profiles:
        names = [profile["current_name"], *json_list(profile["recognition_names_json"])]
        normalized_names = sorted({normalize_name(name) for name in names if normalize_name(name)})
        projects = sorted(set(json_list(profile["recognition_projects_json"])))
        base = {
            "identity_key": profile["identity_key"],
            "current_name": profile["current_name"],
            "recognition_names": "；".join(names),
            "recognition_projects": "；".join(projects),
        }
        if projects and set(projects).issubset(EXEMPT_PROJECTS):
            exempt_only.append({**base, "reason": "项目级身份核验豁免；不得据此伪造信用代码"})

        matches: list[dict[str, str]] = []
        for normalized in normalized_names:
            matches.extend(archived_by_name.get(normalized, []))
        distinct_codes = sorted({item["unified_social_credit_code"] for item in matches})

        if len(distinct_codes) == 1:
            code = distinct_codes[0]
            preferred = next(
                (
                    item
                    for item in matches
                    if normalize_name(item.get("import_name")) in normalized_names
                    and item["unified_social_credit_code"] == code
                ),
                matches[0],
            )
            unique_candidates.append(
                {
                    **base,
                    "candidate_current_name": preferred.get("current_name", ""),
                    "unified_social_credit_code": code,
                    "matched_value": preferred.get("matched_value", ""),
                    "matched_column": preferred.get("matched_column", ""),
                    "evidence_status": preferred["evidence_status"],
                    "source_file": preferred["source_file"],
                    "source_sheet": preferred["source_sheet"],
                    "archive_match_count": len(matches),
                }
            )
            coverage_by_project.update(projects)
        elif len(distinct_codes) > 1:
            conflicts.append(
                {
                    **base,
                    "candidate_codes": "；".join(distinct_codes),
                    "candidate_names": "；".join(sorted({item.get("current_name", "") for item in matches if item.get("current_name")})),
                    "source_files": "；".join(sorted({item["source_file"] for item in matches})),
                    "reason": "同一现名或认定名命中多个信用代码，禁止自动转正",
                }
            )
            conflict_by_project.update(projects)
        else:
            missing.append({**base, "reason": "现有归档批次未命中有效统一社会信用代码"})
            missing_by_project.update(projects)

    write_csv(
        output_dir / "待核主体_归档批次唯一代码候选.csv",
        [
            "identity_key", "current_name", "recognition_names", "recognition_projects",
            "candidate_current_name", "unified_social_credit_code", "matched_value",
            "matched_column", "evidence_status", "source_file", "source_sheet",
            "archive_match_count",
        ],
        unique_candidates,
    )
    write_csv(
        output_dir / "待核主体_身份代码冲突.csv",
        [
            "identity_key", "current_name", "recognition_names", "recognition_projects",
            "candidate_codes", "candidate_names", "source_files", "reason",
        ],
        conflicts,
    )
    write_csv(
        output_dir / "待核主体_归档真实未命中.csv",
        ["identity_key", "current_name", "recognition_names", "recognition_projects", "reason"],
        missing,
    )
    write_csv(
        output_dir / "项目级免核验主体.csv",
        ["identity_key", "current_name", "recognition_names", "recognition_projects", "reason"],
        exempt_only,
    )
    verified_increment = [
        row
        for row in unique_candidates
        if str(row["unified_social_credit_code"]) in verified_codes
    ]
    second_source_queue = [
        row
        for row in unique_candidates
        if str(row["unified_social_credit_code"]) not in verified_codes
        and not set(str(row["recognition_projects"]).split("；")).issubset(EXEMPT_PROJECTS)
    ]
    first_source_queue = [
        row
        for row in missing
        if not set(str(row["recognition_projects"]).split("；")).issubset(EXEMPT_PROJECTS)
    ]
    write_csv(
        output_dir / "需独立复核的单源代码候选.csv",
        [
            "identity_key", "current_name", "recognition_names", "recognition_projects",
            "candidate_current_name", "unified_social_credit_code", "matched_value",
            "matched_column", "evidence_status", "source_file", "source_sheet",
            "archive_match_count",
        ],
        second_source_queue,
    )
    write_csv(
        output_dir / "需要首次身份采集的主体.csv",
        ["identity_key", "current_name", "recognition_names", "recognition_projects", "reason"],
        first_source_queue,
    )

    report = {
        "schema_version": "zhejiang-pending-identity-evidence-audit-v1",
        "database": str(database),
        "inspect_dir": str(inspect_dir),
        "raw_xlsx_dir": str(raw_xlsx_dir) if raw_xlsx_dir else "",
        "verified_checklist": str(verified_checklist) if verified_checklist else "",
        "archive": archive_stats,
        "pending_profiles": len(pending_profiles),
        "unique_code_candidates": len(unique_candidates),
        "code_conflicts": len(conflicts),
        "archive_missing": len(missing),
        "project_exempt_only": len(exempt_only),
        "verified_increment_profiles": len(verified_increment),
        "verified_increment_codes": len(
            {str(row["unified_social_credit_code"]) for row in verified_increment}
        ),
        "single_source_second_source_queue": len(second_source_queue),
        "first_source_collection_queue": len(first_source_queue),
        "coverage_by_project": dict(sorted(coverage_by_project.items())),
        "conflicts_by_project": dict(sorted(conflict_by_project.items())),
        "missing_by_project": dict(sorted(missing_by_project.items())),
        "promotion_boundary": (
            "唯一代码候选只证明旧会话已采集到商业批次结果；在独立登记源或第二来源复核前，"
            "不得写成knowledge_verified。"
        ),
        "outward_source_label": "共创研究院知识库",
    }
    (output_dir / "待核主体_归档批次覆盖审计.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", required=True, type=Path)
    parser.add_argument("--inspect-dir", required=True, type=Path)
    parser.add_argument("--raw-xlsx-dir", type=Path)
    parser.add_argument("--verified-checklist", type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    report = reconcile(
        args.database,
        args.inspect_dir,
        args.raw_xlsx_dir,
        args.verified_checklist,
        args.output_dir,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
