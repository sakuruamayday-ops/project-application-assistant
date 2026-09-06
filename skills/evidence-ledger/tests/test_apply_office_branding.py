from __future__ import annotations

import json
import subprocess
import sys
import zipfile
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook


SCRIPT = Path(__file__).parents[1] / "scripts" / "apply_office_branding.py"
IDENTITY = "共创研究院"
DOCX_WATERMARK = "Gongchuang Institute Centered Watermark v4"
XLSX_MARKER = "_GONGCHUANG_INSTITUTE_UNIFORM_WATERMARK_V4"


def run(path: Path) -> dict[str, object]:
    completed = subprocess.run(
        [sys.executable, str(SCRIPT), str(path)],
        check=True,
        capture_output=True,
        text=True,
    )
    # 签名运行时要求 stdout 只有一个 JSON 对象。依赖库警告若混入 stdout，
    # 即使文件已修改也必须在发布前失败，避免客户端误以为品牌操作未执行。
    assert completed.stdout.count("\n") == 1
    return json.loads(completed.stdout)


def test_applies_signed_branding_to_docx_once(tmp_path: Path) -> None:
    artifact = tmp_path / "report.docx"
    document = Document()
    document.add_paragraph("GC-QA 正文")
    document.save(artifact)

    first = run(artifact)
    second = run(artifact)

    assert first["schema_version"] == "gongchuang-office-branding-operation/v1"
    assert second["status"] == "passed"
    with zipfile.ZipFile(artifact) as package:
        headers = [
            package.read(name).decode("utf-8")
            for name in package.namelist()
            if name.startswith("word/header") and name.endswith(".xml")
        ]
    assert headers
    assert all(IDENTITY in header for header in headers)
    assert all(header.count(DOCX_WATERMARK) == 2 for header in headers)


def test_applies_signed_branding_to_xlsx(tmp_path: Path) -> None:
    artifact = tmp_path / "report.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "GC-QA 正文"
    workbook.save(artifact)

    result = run(artifact)

    assert result["format"] == "xlsx"
    with zipfile.ZipFile(artifact) as package:
        workbook_xml = package.read("xl/workbook.xml").decode("utf-8")
    assert XLSX_MARKER in workbook_xml
    branded = load_workbook(artifact)
    assert IDENTITY in (branded.active.oddHeader.right.text or "")
